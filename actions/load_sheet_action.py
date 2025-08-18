import os
from base_classes import StepwiseAction, Completed
from openpyxl import load_workbook
import csv
from io import StringIO

class LoadSheetAction(StepwiseAction):
    def __init__(self, session):
        self.session = session
        self.tc = session.utils.tab_completion
        self.tc.set_session(session)

    def start(self, args: list | dict | None = None, content=None) -> Completed:
        filename = None
        if isinstance(args, (list, tuple)) and args:
            filename = " ".join(str(a) for a in args)
        elif isinstance(args, dict):
            filename = args.get('file') or args.get('path')

        if not filename:
            self.tc.run('file_path')
            filename = self.session.ui.ask_text("Enter spreadsheet filename (or q to exit): ")
            if str(filename).lower().strip() == 'q':
                self.tc.run('chat')
                return Completed({'ok': True, 'cancelled': True})

        ok = self._process_sheet(str(filename))
        self.tc.run('chat')
        return Completed({'ok': ok, 'file': str(filename)})

    def resume(self, state_token: str, response) -> Completed:
        if isinstance(response, dict) and 'response' in response:
            response = response['response']
        filename = str(response or '')
        if not filename or filename.lower() == 'q':
            self.tc.run('chat')
            return Completed({'ok': True, 'cancelled': True})
        ok = self._process_sheet(filename)
        self.tc.run('chat')
        return Completed({'ok': ok, 'file': filename, 'resumed': True})

    def _process_sheet(self, file_path):
        try:
            wb = load_workbook(filename=file_path, read_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                csv_output = StringIO()
                csv_writer = csv.writer(csv_output)

                for row in sheet.iter_rows(values_only=True):
                    csv_writer.writerow(row)

                csv_content = csv_output.getvalue()

                self.session.add_context('sheet', {
                    'name': f"{os.path.basename(file_path)} - {sheet_name}",
                    'content': csv_content,
                    'metadata': {'original_file': file_path, 'sheet_name': sheet_name}
                })

                try:
                    self.session.ui.emit('status', {'message': f"Sheet '{sheet_name}' from {file_path} added to context."})
                except Exception:
                    pass

            return True
        except Exception as e:
            try:
                self.session.ui.emit('error', {'message': f"Error processing spreadsheet {file_path}: {str(e)}"})
            except Exception:
                pass
            return False
