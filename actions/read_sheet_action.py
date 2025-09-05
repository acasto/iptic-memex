from __future__ import annotations

import os
import csv
from io import BytesIO, StringIO
from base_classes import InteractionAction

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore


class ReadSheetAction(InteractionAction):
    """Extract CSV per sheet from an XLSX and add them as file contexts."""

    def __init__(self, session):
        self.session = session

    def process(self, path: str, *, fs_handler=None) -> bool:
        try:
            if load_workbook is None:
                raise RuntimeError('openpyxl not available')
            # Read via assistant fs handler when provided; else utils FS
            if fs_handler:
                data = fs_handler.read_file(path, binary=True)
            else:
                abs_path = self.session.utils.fs.resolve_file_path(path, extension='.xlsx') or path
                data = self.session.utils.fs.read_file(abs_path, binary=True)
            if data is None:
                return False

            wb = load_workbook(filename=BytesIO(data), read_only=True)
            base = os.path.basename(path)
            ok_any = False
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                csv_output = StringIO()
                writer = csv.writer(csv_output)
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow(row)
                content = csv_output.getvalue()
                self.session.add_context('file', {
                    'name': f"{base} - {sheet_name}",
                    'content': content,
                    'metadata': {'original_file': path, 'sheet_name': sheet_name}
                })
                ok_any = True
            if ok_any:
                try:
                    self.session.ui.emit('status', {'message': f"Loaded XLSX: {base}"})
                except Exception:
                    pass
            return ok_any
        except Exception:
            try:
                self.session.ui.emit('error', {'message': f"Failed to process XLSX: {path}"})
            except Exception:
                pass
            return False

    def run(self, args=None, content=None):
        pass

