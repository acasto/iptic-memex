from __future__ import annotations

import os
from io import StringIO
from pathlib import Path
from typing import Tuple, Dict, Any, Optional
import warnings


def _safe_import(name: str):
    try:
        return __import__(name)
    except Exception:
        return None


def get_supported_exts() -> set[str]:
    return {".pdf", ".docx", ".xlsx"}


def get_versions() -> Dict[str, str | None]:
    out: Dict[str, str | None] = {}
    # Prefer pypdf; include PyPDF2 only if present (suppress deprecation warnings)
    pypdf = _safe_import('pypdf')
    out['pypdf'] = getattr(pypdf, '__version__', None) if pypdf else None
    with warnings.catch_warnings():
        warnings.filterwarnings('ignore', category=DeprecationWarning, module='PyPDF2')
        PyPDF2 = _safe_import('PyPDF2')
    out['PyPDF2'] = getattr(PyPDF2, '__version__', None) if PyPDF2 else None
    docx = _safe_import('docx')
    out['python-docx'] = getattr(docx, '__version__', None) if docx else None
    openpyxl = _safe_import('openpyxl')
    out['openpyxl'] = getattr(openpyxl, '__version__', None) if openpyxl else None
    return out


def extract_text_for_file(path: str) -> Tuple[str | None, Dict[str, Any]]:
    """Extract plain text for known binary document types.

    Returns (text or None if unsupported/failed, metadata dict).
    """
    ext = Path(path).suffix.lower()
    if ext == '.pdf':
        return _extract_pdf(path)
    if ext == '.docx':
        return _extract_docx(path)
    if ext == '.xlsx':
        return _extract_xlsx(path)
    return None, {'kind': 'unknown'}


def _extract_pdf(path: str) -> Tuple[str | None, Dict[str, Any]]:
    meta: Dict[str, Any] = {'kind': 'pdf'}
    # Prefer pypdf; fall back to PyPDF2 with deprecation suppressed
    PdfReader = None
    try:
        from pypdf import PdfReader as _PdfReader  # type: ignore
        PdfReader = _PdfReader
    except Exception:
        try:
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=DeprecationWarning, module='PyPDF2')
                from PyPDF2 import PdfReader as _PdfReader  # type: ignore
            PdfReader = _PdfReader
        except Exception:
            return None, {**meta, 'error': 'pypdf/PyPDF2 not available'}
    try:
        pdf = PdfReader(path)
        if getattr(pdf, 'is_encrypted', False):
            return None, {**meta, 'error': 'encrypted'}
        pages = getattr(pdf, 'pages', [])
        text_parts = []
        count = 0
        for page in pages:
            try:
                t = page.extract_text() or ''
            except Exception:
                t = ''
            text_parts.append(t)
            count += 1
        meta['pages'] = count
        full = "\n".join(text_parts).strip()
        return (full if full else None), meta
    except Exception as e:
        return None, {**meta, 'error': str(e)}


def _extract_docx(path: str) -> Tuple[str | None, Dict[str, Any]]:
    meta: Dict[str, Any] = {'kind': 'docx'}
    try:
        from docx import Document  # type: ignore
    except Exception:
        return None, {**meta, 'error': 'python-docx not available'}
    try:
        doc = Document(path)
        paras = [p.text or '' for p in doc.paragraphs]
        meta['paragraphs'] = len(paras)
        full = "\n".join(paras).strip()
        return (full if full else None), meta
    except Exception as e:
        return None, {**meta, 'error': str(e)}


def _extract_xlsx(path: str) -> Tuple[str | None, Dict[str, Any]]:
    meta: Dict[str, Any] = {'kind': 'xlsx'}
    try:
        from openpyxl import load_workbook  # type: ignore
        import csv
    except Exception:
        return None, {**meta, 'error': 'openpyxl/csv not available'}
    try:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        sheet_names = list(wb.sheetnames)
        meta['sheets'] = len(sheet_names)
        out = StringIO()
        writer = csv.writer(out)
        for sname in sheet_names:
            try:
                sheet = wb[sname]
            except Exception:
                continue
            # Header to mark sheet boundaries
            out.write(f"# Sheet: {sname}\n")
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
            out.write("\n")
        text = out.getvalue().strip()
        return (text if text else None), meta
    except Exception as e:
        return None, {**meta, 'error': str(e)}
